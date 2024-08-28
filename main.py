# -*- coding: utf8 -*-
from aiogram import Bot, types
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import Dispatcher
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters import Command
from aiogram.dispatcher.filters.state import StatesGroup, State
from aiogram.utils import executor
from bs4 import BeautifulSoup
from datetime import date, datetime, timedelta, timezone
from dotenv import load_dotenv
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl import Workbook
from os import environ as env
from telethon.sync import TelegramClient
from tp_stat_bot import bot_keyboard, keep_alive, get_vk_stats, get_visionboard
import googleapiclient.discovery
import httplib2
import json
import logging
import os, sys
import re
import requests
import statistics

load_dotenv()

PHONE = env['PHONE']
API_ID = int(env['API_ID'])
API_HASH = env['API_HASH']
BOT_KEY = env['BOT_KEY']
SPREADSHEET_ID = '1DYH6LNB9cNVbFHA4593ZbEZl73yCb9_8oYfoeF4QqxM'
URL_PREFIX = 'https://t.me/'
MAX_POST_LENGTH = 4096
AD_STOPWORDS = ["Реклама", "#партнерский", "#партнёрский", "#ивент", "#вакансии", "#вакансия", "вакансии", "Рекламодатель", "#конкурс"]
CREDENTIALS_FILE = 'creds.json'
TG_CHANNELS_COUNT = 27

with open('channels.json') as channels_obj:
  channels = json.load(channels_obj)

keep_alive.keep_alive()

storage = MemoryStorage()  # Для машины конечных состояний (FSM)
bot = Bot(token=BOT_KEY, parse_mode=types.ParseMode.HTML)
dp = Dispatcher(bot, storage=storage)

class UserState(StatesGroup):
  name = State()
  sum = State()
  statistic = State()
  archive = State()
  vision_board = State()

# TODO: Вместо этого можно использовать loguru, там просто from loguru import logger и красота
logging.basicConfig(
  format=
  u'%(filename)s [LINE:%(lineno)d] #%(levelname)-8s [%(asctime)s]  %(message)s',
  level=logging.INFO,
)

client = TelegramClient(PHONE, API_ID, API_HASH)
client.start()

@dp.message_handler(Command("start"), state=None)
async def welcome(message):
  await bot.send_message(
    message.chat.id,
    f"Привет, {message.from_user.first_name}! Нажмите кнопку «Собрать статистику», чтобы начать :)",
    reply_markup=bot_keyboard.start,
    parse_mode='Markdown')

@dp.message_handler(content_types=['text'])
async def user_register(msg: types.Message):
  if msg.text == "Собрать статистику":
    text = get_channels_list(channels)
    await bot.send_message(msg.from_user.id,
                           text,
                           reply_markup=bot_keyboard.start)
    await UserState.statistic.set()

  elif msg.text == "Суммаризация статьи":
    text = "Вставьте ссылку на статью:\n\n"
    await bot.send_message(msg.from_user.id,
                           text,
                           reply_markup=bot_keyboard.start)
    await UserState.sum.set()

  elif msg.text == "Собрать архив постов":
    text = get_channels_list(channels) + "\nЧтобы собрать архив сразу по всем Telegram-каналам, введите 99."
    await bot.send_message(msg.from_user.id,
                           text,
                           reply_markup=bot_keyboard.start)
    await UserState.archive.set()

  elif msg.text == "Составить вижнборд":
    text = get_channels_list(channels)
    await bot.send_message(msg.from_user.id,
                           text,
                           reply_markup=bot_keyboard.start)
    await UserState.vision_board.set()
  else:
    await msg.answer("Cначала нажмите одну из кнопок")

@dp.message_handler(state=UserState.archive)
async def get_username(msg: types.Message, state: FSMContext):
  await state.update_data(channel=msg.text)
  data = await state.get_data()
  print(str(data))
  user_choice = 0

  try:
    user_choice = int(data['channel'])
    print(str(user_choice))
    print(str(channels))
    if user_choice == 99:
      i = 0
      for i in range(0, TG_CHANNELS_COUNT):

        chat_name_url = [channels[i]['channel_title'], channels[i]['channel_username']]
        print('\nchat_name_url: ', chat_name_url)

        chat = URL_PREFIX + channels[i]['channel_username']
        print('\nchat: ', chat)
        
        await get_posts_archive(chat, chat_name_url, msg)

    elif user_choice > 0 and user_choice <= TG_CHANNELS_COUNT:
      chat_name_url = [channels[user_choice - 1]['channel_title'], channels[user_choice - 1]['channel_username']]
      chat = URL_PREFIX + chat_name_url[1]
      await get_posts_archive(chat, chat_name_url, msg)

    else:
      await push_button_again("«Собрать архив постов»", msg.from_user.id, e)

  except Exception as e:
    await push_button_again("«Собрать архив постов»", msg.from_user.id, e)
  await state.finish()

@dp.message_handler(state=UserState.sum)
async def get_username(msg: types.Message, state: FSMContext):
  try:
    endpoint = 'https://300.ya.ru/api/sharing-url'
    response = requests.post(endpoint,
                             json = {'article_url': msg.text},
                             headers = {'Authorization': "your_key"})
    summaryText = response.json()
    req = requests.get(summaryText['sharing_url'])
    req.encoding = 'utf-8'
    data_site = BeautifulSoup(req.text, 'html.parser')
    text = ""
    data1 = data_site.find('ul')

    for li in data1.find_all("li"):
        text += str(li.text)[3:]

    text += 'Ссылка на текст: ' + summaryText['sharing_url']
    await bot.send_message(msg.from_user.id, str(text))
    await state.finish()
  
  except Exception as e:
    await push_button_again("«Суммаризация статьи»", msg.from_user.id, e)
    await state.finish()

@dp.message_handler(state=UserState.statistic)
async def get_statistics(msg: types.Message, state: FSMContext):
  await state.update_data(statistic=msg.text)
  data = await state.get_data()
  chat = ""
  if 0 < int(data['statistic']) <= TG_CHANNELS_COUNT:
    try:
      chat = channels[int(data['statistic']) - 1]
      chat = URL_PREFIX + chat['channel_username']
      # Определяем период сбора статистики в зависимости от дня недели
      startDateWeek, endDateWeek = get_period_week()
      # Отправляем сообщение о том, как собирается статистика
      await send_pre_statistic_message(msg, startDateWeek, endDateWeek, chat)

      # Получаем строку с реакциями для каждого поста
      stringPostsStat, allReactions, dict_posts_er = await get_week_reactions_and_views(chat, startDateWeek, endDateWeek)
      # Получаем значения для финальной статистики
      valuesForStatistic = await get_all_values_for_statistics(chat, allReactions)
      # Получаем текст с финальной статистикой и значениями для вставки
      finalStatText = get_summary_message_text(*valuesForStatistic)

      # Отправляем сообщение с реакциями по каждому посту (режем, если оно больше 4096 символов)
      await cut_and_send_stat_message(stringPostsStat, msg.from_user.id)
      # Отправляем сообщение с финальной статистикой и значениями для вставки (режем, если он больше 4096 символов)
      await cut_and_send_stat_message(finalStatText, msg.from_user.id)
      
      # Заканчиваем состояние
      await state.finish()

    except Exception as e:
      await push_button_again("«Собрать статистику»", msg.from_user.id, e)
      await state.finish()
  else:
    chat = channels[int(data['statistic']) - 1]
    chat = 'https://vk.com/' + chat['domain']

    domain = channels[int(data['statistic']) - 1]['domain']
    ownerId = channels[int(data['statistic']) - 1]['owner_id']
  
    startDateWeek, endDateWeek = get_period_week()
    await send_pre_statistic_message(msg, startDateWeek, endDateWeek, chat)

    await bot.send_message(msg.from_user.id, get_vk_stats.get_metrics(domain, ownerId),
                        reply_markup=bot_keyboard.start)
    await state.finish()

@dp.message_handler(state=UserState.vision_board)
async def get_vision_board(msg: types.Message, state: FSMContext):
  try:
    await state.update_data(statistic=msg.text)
    data = await state.get_data()
    chat = ""
    if 0 < int(data['statistic']) <= TG_CHANNELS_COUNT:
      chat = channels[int(data['statistic']) - 1]
      chat = URL_PREFIX + chat['channel_username']
      # Определяем период сбора статистики в зависимости от дня недели
      start_date, end_date = get_period_month()

      # Отправляем сообщение о том, как собирается статистика для вижнборда
      await get_visionboard.send_pre_visionboard_message(bot, msg, start_date, end_date, chat)

      # Получаем строку с реакциями для каждого поста
      stringPostsStat, allReactions, dict_posts_er = await get_week_reactions_and_views(chat, start_date, end_date)
      print('allReactions = ', allReactions)

      # Получаем среднее количество всех реакций
      # TODO Заменить на стандартные значения для каждого канала
      er_s = statistics.median(allReactions)
      # Получаем 4 списка постов по ER
      er_worst, er_bad, er_good, er_great = await get_visionboard.sort_vision_board(dict_posts_er, er_s)

      # Отправляем сообщение с подробными реакциями по каждому посту
      await get_visionboard.send_all_reactions(bot, msg, dict_posts_er, er_worst, er_bad, er_good, er_great)
      
      # Создаем картинку вижнборда
      output_visionboard = await get_visionboard.create_visionboard(er_worst, er_bad, er_good, er_great)
      # Отправляем вижнборд
      print('\nОтправляю вижнборд и рекомендации в бота.')
      f = open(output_visionboard, 'rb')
      await bot.send_document(msg.from_user.id, f)
      # Отправляем инструкцию по работе с вижнбордом
      await get_visionboard.send_after_visionboard_message(bot, msg)

      await state.finish()
    else:
      await bot.send_message(msg.from_user.id, "Для выбранного канала пока невозможно составить вижноборд. Попробуйте позже.", disable_web_page_preview=True)
      await state.finish()
  except Exception as e:
      await push_button_again("«Составить вижнборд»", msg.from_user.id, e)
      await state.finish()

async def get_week_reactions_and_views(chat, start_date, end_date):
  stringPostsStat = ""
  countViews = []
  allReactions = []
  dict_posts_er = {}
  prvGroupID = 0

  async for message in client.iter_messages(chat, reverse=True, offset_date=start_date):
    print("\nПроверка поста: " + chat + "/" + str(message.id))
    if (message.date <= end_date):
      print("— Подходит по дате.")

      # Если слова есть в посте, значит, это реклама, выходим из итерации
      if any(word in message.message for word in AD_STOPWORDS):
        print("— Есть реклама, не считаем.")

        # Проверяем, есть ли у поста медиа и grouped_id, если да, дальше будет группа картинок
        if (message.media and message.grouped_id):
          if prvGroupID == message.grouped_id: continue
          else: prvGroupID = message.grouped_id
        continue

      # Проверяем, есть ли у поста медиа и grouped_id, если да, дальше будет группа картинок
      print("— Нет рекламы, считаем.")
      if (message.media and message.grouped_id):
        if prvGroupID == message.grouped_id:
          print("— Есть grouped_id, не считаем: " + str(prvGroupID))
          continue
        else:
          prvGroupID = message.grouped_id
          print("— Есть grouped_id, но пост первый, считаем: " + str(prvGroupID))
      print("— Нет grouped_id, считаем")

      # Проверяем сообщение на None (например, закреп)
      if message.views is None:
        print("— Технический пост, не считаем.")
        continue
      else:
        print("— Реальный пост, считаем.")
        countViews.append(int(message.views))

      print("— Сбор реакций запущен.")

      # Забираем комменты, реакции на комментах
      countComOnP, countReactOnKFinal = await get_comments(message, chat, message.id)

      # Считаем сумму всех реакций
      countAllReactMP = get_sum_all_posts_reactions(message, countComOnP, countReactOnKFinal)

      # Собираем информацию по каждому посту в строку
      string_posts_stat_temporary = get_string_stat_for_post(message, chat, countComOnP, countReactOnKFinal, countAllReactMP)
      stringPostsStat += string_posts_stat_temporary

      # Добавляем сумму реакций каждого поста в список

      allReactions.append(int(countAllReactMP))

      # TODO: Убрать или доработать исключение, оно не верное
      try:
        adr_post = chat + "/" + str(message.id)
        dict_posts_er.setdefault(adr_post, [])

        print('— Вставляю в словарь сумму всех реакций: ', countAllReactMP)
        dict_posts_er[adr_post].append(countAllReactMP)
        print('— Вставляю в словарь строку с реакциями: ', string_posts_stat_temporary)
        dict_posts_er[adr_post].append(string_posts_stat_temporary)
        print("— Сумма и строка реакций записаны в словарь по адресу поста: " + str(dict_posts_er[adr_post]))
      except Exception as e: print(e)

      print("— Сбор реакций завершен.")
  return stringPostsStat, allReactions, dict_posts_er

async def get_avg_views_for_month(chat, median=False):
  # Присваиваем даты для подсчёта постов за месяц
  start_date, end_date = get_month_for_stats()
  countViews = []
  prvGroupID = 0
  print("\n\nПолучаем средний и медианный охват поста за месяц:")
  async for message in client.iter_messages(chat, reverse=True, offset_date=start_date):
  
    if (message.date <= end_date):
      print("Пост: " + chat + "/" + str(message.id) + " : " + str(message.date))
      # Проверяем, есть ли у поста медиа и grouped_id, если да, дальше будет группа картинок
      if (message.media and message.grouped_id):
        if prvGroupID == message.grouped_id: continue
        else: prvGroupID = message.grouped_id
        # Проверяем сообщение на None (например, закреп)
      if message.views is None: continue
      else: countViews.append(int(message.views))
  # Получаем средние просмотры на пост
  avgMonthViews = sum(list(map(int, countViews))) // len(countViews)
  text = "\nСредний охват поста за месяц получен: "

  if median == True:
    # Получаем медианные просмотры на пост
    avgMonthViews = statistics.median(list(map(int, countViews)))
    text = "\nМедианный охват поста за месяц получен: "

  print(text + str(avgMonthViews))
  return avgMonthViews

async def send_pre_statistic_message(msg, start_date, end_date, chat):
  msgDateInfo = ("<b>Выбранный канал: " +
                 chat + "</b>\n\n— Собираем статистику по постам от понедельника <b>" +
                 str(start_date.strftime('%d.%m.%Y')) + "</b> (включительно) и до понедельника <b>" +
                 str(end_date.strftime('%d.%m.%Y')) + "</b> (не включительно)." +
                 "\n— Подписки и отписки собираются со сдвигом в три дня вперёд: со среды по среду." +
                 "\n— Кол-во подписчиков — во вторник." +
                 "\n— Cтатистика за предыдущую неделю собирается только со среды текущей недели." +
                 " В понедельник и вторник будет старая статистика." +
                 "\n\n<b>Бот начал собирать статистику. Это займёт не больше минуты.</b>")
  await bot.send_message(msg.from_user.id, msgDateInfo, disable_web_page_preview=True)

async def send_vk_stats_message(msg, start_date, end_date, chat):
  msgDateInfo = ("<b>Выбранный канал: " +
                 chat + "</b>\n\n— Собираем статистику по постам от понедельника <b>" +
                 str(start_date.strftime('%d.%m.%Y')) + "</b> (включительно) и до понедельника <b>" +
                 str(end_date.strftime('%d.%m.%Y')) + "</b> (не включительно)." +
                 "\n— Подписки и отписки собираются со сдвигом в три дня вперёд: со среды по среду." +
                 "\n— Кол-во подписчиков — во вторник." +
                 "\n— Cтатистика за предыдущую неделю собирается только со среды текущей недели." +
                 " В понедельник и вторник будет старая статистика." +
                 "\n\n<b>Бот начал собирать статистику. Это займёт не больше минуты.</b>")
  await bot.send_message(msg.from_user.id, msgDateInfo, disable_web_page_preview=True)

def get_month_for_stats():

  # Выберем даты прошлого месяца
  end_date = datetime.now().replace(day = 1) - timedelta(days = 1)
  start_date = datetime.now().replace(day = 1) - timedelta(days = end_date.day)

  # Обнуляем значения часов, минут и секунд
  start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
  end_date = end_date.replace(hour=0, minute=0, second=0, microsecond=0, tzinfo=timezone.utc)
  return start_date, end_date

async def get_subscriber_timeline(chat):
  # Получаем общую статистику
  channelFullInfo = await client.get_stats(chat)
  # Получаем графики подписок / отписок
  subscriptionsUnsubscribesGraph = str(channelFullInfo.followers_graph)
  # Получаем графики общего числа подписчиков
  allSubscribersGraph = str(channelFullInfo.growth_graph)
  return subscriptionsUnsubscribesGraph, allSubscribersGraph

def get_unsubscribes(endidxOfMemberList, startidxOfMemberList, subscriptionsUnsubscribesGraph):
  # Забираем столбец с отписками
  start = subscriptionsUnsubscribesGraph.find('"y1",') + 5
  end = subscriptionsUnsubscribesGraph.find(']', start)
  unsubMembersList = [int(i) for i in subscriptionsUnsubscribesGraph[start:end].split(',')]
  sumJoinedMembers = sum(
    list(
      map(int, unsubMembersList[endidxOfMemberList:startidxOfMemberList])))
  return sumJoinedMembers

def get_subscribes(endidxOfMemberList, startidxOfMemberList, subscriptionsUnsubscribesGraph):
  # Получаем столбец с подписками
  start = subscriptionsUnsubscribesGraph.find('"y0",') + 5
  end = subscriptionsUnsubscribesGraph.find(']', start)

  joinedMembersList = [int(i) for i in subscriptionsUnsubscribesGraph[start:end].split(',')]
  sumJoinedMembers = list(
    map(int, joinedMembersList[endidxOfMemberList:startidxOfMemberList]))
  sumJoinedMembers = sum(sumJoinedMembers)
  return sumJoinedMembers

def get_number_of_subscribers(endidxOfMemberList, startidxOfMemberList, allSubscribersGraph):
  # Получаем столбец с общим числом подписчиков
  start = allSubscribersGraph.find('"y0",') + 5
  end = allSubscribersGraph.find(']', start)

  allMembersList = [int(i) for i in allSubscribersGraph[start:end].split(',')]
  numberOfSubscribers = list(
    map(int, allMembersList[endidxOfMemberList:startidxOfMemberList]))
  numberOfSubscribers = numberOfSubscribers[-1]
  return numberOfSubscribers

def get_sum_all_posts_reactions(message, countComOnP, countReactOnKFinal):
  countAllReactMP = (int(get_reposts(message)) + int(get_reactions(message)) +
                     int(get_negative_inline_reactions(message)) +
                     int(get_positive_inline_reactions(message)) +
                     int(countComOnP) +
                     int(countReactOnKFinal) +
                     int(get_poll_answers(message)))
  return countAllReactMP

def get_string_stat_for_post(message, chat, countComOnP, countReactOnKFinal, countAllReactMP):
  stringStatForPost = (chat + "/" + str(message.id) + "\n" +
                          str(get_reposts(message)) + "📢 + " +
                          str(get_reactions(message)) + "❤️ + " +
                          str(int(get_positive_inline_reactions(message)) + int(get_negative_inline_reactions(message))) + "👍👎 + " +
                          str(countComOnP) + "💬 + " +
                          str(countReactOnKFinal) + "❤️💬 + " +
                          str(get_poll_answers(message)) + "📊 = " +
                          str(countAllReactMP)+ "\n\n")
  return stringStatForPost

async def get_all_values_for_statistics(chat, allReactions):
  # Получаем графики подписок/отписок и общего числа подписчиков
  subscriptionsUnsubscribesGraph, allSubscribersGraph = await get_subscriber_timeline(chat)
  # Получаем даты для списка подписок/отписок
  startidxOfMemberList, endidxOfMemberList = get_idx_of_member_list()

  # Отписки:
  sumUnsubMembers = get_unsubscribes(endidxOfMemberList, startidxOfMemberList, subscriptionsUnsubscribesGraph)
  # Подписки:
  sumJoinedMembers = get_subscribes(endidxOfMemberList, startidxOfMemberList, subscriptionsUnsubscribesGraph)
  # Количество подписчиков:
  numberOfSubscribers = get_number_of_subscribers(endidxOfMemberList, startidxOfMemberList, allSubscribersGraph)
  # Отписок в день:
  procUnsubDay = round((sumUnsubMembers / numberOfSubscribers / 7) * 100, 2)
  # Подписок в день:
  procJoinedDay = round((sumJoinedMembers / numberOfSubscribers / 7) * 100, 2)
  # Средний охват поста:
  avgMonthViews = await get_avg_views_for_month(chat)
  # Медианный охват поста:
  median_month_views = await get_avg_views_for_month(chat, median=True)
  # Доля активной аудитории:
  activeAudit = round((avgMonthViews / numberOfSubscribers) * 100, 2)
  # Медиана реакий:
  reactionsMedian = statistics.median(allReactions)
  # Доля вовлеченной аудитории:
  engagedAudit = round((statistics.median(allReactions) / avgMonthViews) * 100, 2)

  return sumUnsubMembers, sumJoinedMembers, procUnsubDay, procJoinedDay, numberOfSubscribers, avgMonthViews, median_month_views, activeAudit, reactionsMedian, engagedAudit

def get_summary_message_text(sumUnsubMembers, sumJoinedMembers, procUnsubDay, procJoinedDay, numberOfSubscribers, avgMonthViews, median_month_views, activeAudit, reactionsMedian, engagedAudit):
  text = ("<b>Итоговая статистика</b>\n\n" +
                     "<b>Отписки: </b>" + str(sumUnsubMembers) +
                     "\n<b>Подписки: </b>" + str(sumJoinedMembers) +
                     "\n<b>Отписок в день: </b>" + str(procUnsubDay) + "%" +
                     "\n<b>Подписок в день: </b>" + str(procJoinedDay) + "%" +
                     "\n<b>Количество подписчиков: </b>" + str(numberOfSubscribers) +
                     "\n<b>Средний охват поста за месяц: </b>" + str(avgMonthViews) + f' (Медиана: {median_month_views})'
                     "\n<b>Доля активной аудитории: </b>" + str(activeAudit) + "%" +
                     "\n<b>Медиана реакций: </b>" + str(reactionsMedian) +
                     "\n<b>Доля вовлечённой аудтиории: </b>" + str(engagedAudit) + "%" +

                     "\n\n<b>Значения для вставки:</b>\n<code>" +
                     str(sumUnsubMembers) + "\n" +
                     str(sumJoinedMembers) + "\n" +
                     str(procUnsubDay) + "%" + "\n" +
                     str(procJoinedDay) + "%" + "\n" +
                     str(numberOfSubscribers) + "\n" +
                     str(avgMonthViews) + "\n" +
                     str(activeAudit) + "%" + "\n" +
                     str(reactionsMedian) + "\n" +
                     str(engagedAudit) + "%\n" +
                     str(median_month_views) + "</code>"

                     "\n\nПроверьте, что в постах совпадают даты, реакции и нет рекламы." +
                     "\n\n<i>Если бот помог вам, не забудьте мысленно отправить плюсик в карму Паше Ф. ❤️</i>")
  return text

def get_vk_stats_text(activeAudienceRate, engagedAudienceRate, reachMedian, reactionsMedian, subscribers):
  text = ("<b>Итоговая статистика</b>\n\n" +
          "<b>Количество подписчиков: </b>" + str(subscribers) +
          "\n<b>Медианный охват контентного поста: </b>" + str(reachMedian) +
          "\n<b>Доля активной аудитории: </b>" + str(activeAudienceRate) + "%" +
          "\n<b>Медианное количество реакций: </b>" + str(reactionsMedian) + 
          "\n<b>Доля вовлечённой аудитории: </b>" + str(engagedAudienceRate) + "%" +
          
          "\n\n<b>Значения для вставки:</b>\n<code>" +
          str(subscribers) + "\n" +
          str(reachMedian) + "\n" +
          str(activeAudienceRate) + "%" + "\n" +
          str(reactionsMedian) + "%" + "\n" +
          str(engagedAudienceRate) + "\n"
          
          "\n\n<i>Если бот помог вам, не забудьте мысленно отправить плюсик в карму Паше Ф. и Лене К. ❤️</i>")
  return text

def get_channels_list(channels):
    i = 0
    channelsList = "<b>Введите номер канала (только цифру):</b>\n\n"

    while i < len(channels):
      channelsList += str(channels[i]['id']) + ".  " + channels[i]['channel_title'] + "\n"
      i += 1

    return channelsList

def get_period_week():
  # Определяем сегодняшний день недели
  weekday = int(datetime.today().isoweekday())
  # Считаем стату со среды предыдущей недели до среды этой
  x = 7 if weekday >= 3 else 14
  # Определяем даты
  start_date = datetime.now()-timedelta(days=datetime.today().weekday()+x)
  end_date = datetime.now()-timedelta(days=datetime.today().weekday()+x-7)
  # Присваиваем таймзону, чтобы не было проблем при сравнении с message.date
  start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
  end_date = end_date.replace(hour=0,minute=0,second=0,microsecond=0,tzinfo=timezone.utc)
  return (start_date, end_date)

def get_period_month():
  # Находим начало и конец предыдущего месяца
  end_date = datetime.today().replace(day=1) - timedelta(days=1)
  start_date = datetime.today().replace(day=1) - timedelta(days=end_date.day)
  # Меняем конечную дату на первый день этого месяца
  end_date = datetime.today().replace(day=1)
  # Присваиваем таймзону, чтобы не было проблем при сравнении с message.date
  start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
  end_date = end_date.replace(hour=0,minute=0,second=0,microsecond=0,tzinfo=timezone.utc)
  print(f'Дата начала сбора архива: {start_date}')
  print(f'Дата конца сбора архива: {end_date}')

  return (start_date, end_date)

def get_idx_of_member_list():
  weekday = int(datetime.today().isoweekday())
  startidxOfMemberList = 0
  endidxOfMemberList = 0
  # Со среды у нас всё ок со счётом. А вот во вт-пн надо считать прерыдущую неделю
  if weekday >= 3:
    startidxOfMemberList = 0 - (weekday - 2)
    endidxOfMemberList = startidxOfMemberList - 7
  else:
    startidxOfMemberList = 0 - (weekday + 5)
    endidxOfMemberList = startidxOfMemberList - 7
  return startidxOfMemberList, endidxOfMemberList

async def cut_and_send_stat_message(stringPostsStat, id):
  if len(stringPostsStat) > MAX_POST_LENGTH:
    for x in range(0, len(stringPostsStat), MAX_POST_LENGTH):
      await bot.send_message(id, stringPostsStat[x:x + MAX_POST_LENGTH], reply_markup=bot_keyboard.start, disable_web_page_preview=True)
  else:
    await bot.send_message(id, stringPostsStat, reply_markup=bot_keyboard.start, disable_web_page_preview=True)

async def get_posts_archive(chat, chat_name_url, msg):
  print('Начинаю сбор архива постов для канала: ' + str(chat_name_url))
  prvGroupID = 0
  # Определяем даты начала и конца месяцчного периода для собра архива
  start_date, end_date = get_period_month()
  print(f'Определил даты: {str(start_date)} — {str(end_date)}')
  # Объявляем список для промежуточного сбора значений для одного листа, чтобы вставить их за один запрос
  excelPasteData = []
  # Создаем эксель для вставки значений
  ws, wb = create_excel()
  # Отправляем сообщение о старте сбора
  await bot.send_message(
    msg.from_user.id,
    "Бот начал собирать посты канала " + chat + " для архива. Это может занять 2-3 минуты на канал.")
  
  # Собираем статистику для архива
  try:
    postsQty = 2
    # Авторизуемся и получаем service — экземпляр доступа к API
    service = get_Google_Api_service()
    print(f'Получил эксземпляр доступа к API гуглодоков: {str(service)}')
    
    # Определяем имя нужного листа
    sheet_name = await get_list_name_google_docs(chat)
    print(f'Название листа для вставки архива: {str(sheet_name)}')
    
    # Определяем последнюю строку и дату последнего поста
    lastRowId, lastRowDate = get_last_row_Google_docs(sheet_name, service, SPREADSHEET_ID)
    print(f'Номер последней строки и данные в ней: {str(lastRowId)} — {str(lastRowDate)}')

    # Определяем сегодняшнюю дату месяц назад, чтобы сравнить с датой последнего поста
    lastMonthDay = get_today_month_ago()

    # Если с даты последнего поста прошло меньше месяца, не собираем посты
    if lastRowDate > lastMonthDay:
      errorMsg = ("Сбор архива постов не удался. С даты последнего поста (" +
                  str(lastRowDate) + ") еще не прошел месяц." +
                  "\nЕсли в архиве Google Docs нет постов, вышедших в прошлом месяце, напишите Паше Ф. о проблеме.")
      await bot.send_message(msg.from_user.id,
                           errorMsg,
                           reply_markup=bot_keyboard.start)
      return

    async for message in client.iter_messages(chat,
                                              reverse=True,
                                              offset_date=start_date):
      if (message.date <= end_date):
        # Проверяем, есть ли у поста медиа и grouped_id, если да, дальше будет группа картинок
        if (message.media and message.grouped_id):
          if prvGroupID == message.grouped_id: continue
          else: prvGroupID = message.grouped_id

        # Забираем дату поста, превью текста, просмотры, репосты, реакции, опросы, реакции на комментах, тип поста
        postsDateExl = str(message.date.replace(tzinfo=None))
        messageTextExl = get_text_preview(message)
        viewsExl = message.views
        countForOnP = get_reposts(message)
        countReactOnP = get_reactions(message)
        countVotes = get_poll_answers(message)
        countComOnP, countReactOnKFinal = await get_comments(message, chat, message.id)
        adsOrContent = get_posts_type(message)
        # Складываем общее количество inline-реакций на кнопках для архива
        likeDislikeExcel = get_positive_inline_reactions(message) + get_negative_inline_reactions(message)

        post_id = chat + '/' + str(message.id)
        print(f'Вставляю данные поста {post_id} в эксель {sheet_name}')
        # Вставляем строку в excel
        paste_to_Excel(ws, postsQty, message.id, postsDateExl, messageTextExl, chat, countForOnP, countReactOnP, countComOnP, countReactOnKFinal, likeDislikeExcel, countVotes, viewsExl, adsOrContent)
        # Вставляем строку со всеми значениями в промежуточный список для гуглодоков
        excelPasteData.append([postsDateExl, messageTextExl, chat + "/" + str(message.id), str(countForOnP), str(countReactOnP), str(countComOnP), str(countReactOnKFinal), likeDislikeExcel, str(countVotes), viewsExl, adsOrContent])
      postsQty += 1
    
    # Записываем в файл
    await get_data_to_google_docs(service, SPREADSHEET_ID, sheet_name, lastRowId, postsQty, excelPasteData)
    # Сохраняем эксель и отправляем пользователю
    await save_Excel(msg, chat_name_url, start_date, end_date, wb)
  
  except Exception as e:
    await push_button_again("«Собрать архив постов»", msg.from_user.id, e)

async def push_button_again(buttonName, id, e):
  errorMsg = ("Попробуйте заново. Нажмите " + buttonName +
              "\n\nОшибка: " + str(e) +
              "\nError on line {}".format(sys.exc_info()[-1].tb_lineno))
  await bot.send_message(id, errorMsg, reply_markup=bot_keyboard.start)

def get_Google_Api_service():
  # Файл, полученный в Google Developer Console
  credentials = ServiceAccountCredentials.from_json_keyfile_name(CREDENTIALS_FILE, ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive'])
  httpAuth = credentials.authorize(httplib2.Http())
  service = googleapiclient.discovery.build('sheets', 'v4', http = httpAuth, cache_discovery=False)
  return service

def get_last_row_Google_docs(sheet_name, service, SPREADSHEET_ID):
  rangeExcel = sheet_name + "!A1:A"
  rows = service.spreadsheets().values().get(spreadsheetId=SPREADSHEET_ID, range=rangeExcel).execute().get('values', [])
  last_row = rows[-1] if rows else None
  lastRowId = len(rows)
  lastRowDate = datetime.strptime(str(last_row[0]), '%Y-%m-%d %H:%M:%S')
  return lastRowId, lastRowDate

async def get_list_name_google_docs(chat):
  id_sheet_by_chats = {"https://t.me/tproger_official": "ТГ ТП",
         "https://t.me/ithumor": "ТГ Юмор",
         "https://t.me/your_tech": "Представляешь,",
         "https://t.me/tproger_web": "ТГ Веб",
         "https://t.me/zen_of_python": "Zen of Python",
         "https://t.me/mobi_dev": "Мобильная разработка",
         "https://t.me/prog_point": "Точка входа",
         "https://t.me/make_game": "GameDev",
         "https://t.me/soft_skillz": "Soft Skillz",
         "https://t.me/neuro_channel": "Нейроканал",
         "https://t.me/a_cup_of_java": "Чашечка Java",
         "https://t.me/prog_stuff": "Сохранёнки программиста",
         "https://t.me/quiz_python": "Python: Задачки и вопросы",
         "https://t.me/prog_tools": "Инструменты программиста",
         "https://t.me/devo_pes": "DevOps для ДевоПсов",
         "https://t.me/typical_qa": "Типичный QA",
         "https://t.me/go_in_action": "Go in Action",
         "https://t.me/linux_n_linus": "Linux и Линус",
         "https://t.me/big_data_analysis": "Data Analysis / Big Data",
         "https://t.me/database_design": "DATABASE DESIGN",
         "https://t.me/django_prog": "Django Unleashed Framework",
         "https://t.me/dot_net_c_sharp": ".NET / C#",
         "https://t.me/smlttech": "SMLTECH"}
  sheet_id = id_sheet_by_chats[chat]
  return sheet_id

def get_today_month_ago():
  today = datetime.today()
  first = today.replace(day=1)
  lastMonth = first - timedelta(days=1)
  lastMonthDay = lastMonth.replace(day=today.day)
  return lastMonthDay

def create_excel():
  # Создаем книгу 
  wb = Workbook()
  # Делаем единственный лист активным 
  ws = wb.active
  # Вставляем названия колонок
  ws['A1'], ws['B1'], ws['C1'], ws['D1'], ws['E1'], ws['F1'], ws['G1'], ws['H1'], ws['I1'], ws['J1'], ws['K1'] = "Дата", "Запись", "Ссылка", "Репосты", "Лайки", "Комментарии", "Лайки на комментарии", "Inline-реакции" , "Голоса", "Просмотры" , "Тип контента"
  return ws, wb

def get_text_preview(message):
  messageText = message.message
  messageText = messageText.replace('\u200b','')
  messageTextExl = messageText.split("\n")[0] + '...'
  return messageTextExl

def get_reactions(message):
  reactStr = str(message.reactions)
  regexNum = re.compile('count=\d+,')
  numReactStr = ""
  regexCount = re.compile('\d+')
  countReactOnP = regexCount.findall(numReactStr.join(regexNum.findall(reactStr)))
  countReactOnP = list(map(int, countReactOnP))
  countReactOnP = sum(countReactOnP)
  return(countReactOnP)

def get_positive_inline_reactions(message):
  reactStr = str(message)
  regexNum = re.compile("text='👍 \d+")
  numForStr = ""
  regexCount = re.compile('\d+')
  countButLOnP = regexCount.findall(numForStr.join(regexNum.findall(reactStr)))
  countButLOnP = list(map(int, countButLOnP))
  countButLOnP = sum(countButLOnP)
  return(countButLOnP)

def get_negative_inline_reactions(message):
  reactStr = str(message)
  regexNum = re.compile("text='👎 \d+")
  numForStr = ""
  regexCount = re.compile('\d+')
  countButDLOnP = regexCount.findall(numForStr.join(regexNum.findall(reactStr)))
  countButDLOnP = list(map(int, countButDLOnP))
  countButDLOnP = sum(countButDLOnP)
  return(countButDLOnP)

def get_poll_answers(message):
  try:
    countVotes = int(message.media.results.total_voters)
  except:
    countVotes = 0
  return countVotes

def get_reposts(message):
  reactStr = str(message)
  regexNum = re.compile('forwards=\d+,')
  numForStr = ""
  regexCount = re.compile('\d+')
  countForOnP = regexCount.findall(numForStr.join(regexNum.findall(reactStr)))
  countForOnP = list(map(int, countForOnP))
  countForOnP = sum(countForOnP)
  return(countForOnP)

async def get_comments(message, chat, i):
  countComOnP = 0
  countReactOnKFinal = 0
  try:
    async for message in client.iter_messages(chat, reply_to=i, reverse=True):
      countComOnP += 1
      if message.reactions != None:
        countReactOnKFinal += get_reactions_on_comment(message)
  except Exception as e:
    print("Не получилось собрать комменты и лайки на них. Пост:" + chat + "/" + str(i) + ". Ошибка: " + str(e))
  return countComOnP, countReactOnKFinal

def get_reactions_on_comment(message):
  reactStr = str(message.reactions)
  regexNum = re.compile('count=\d+,')
  numReactStr = ""
  regexCount = re.compile('\d+')
  countReactOnK = regexCount.findall(numReactStr.join(regexNum.findall(reactStr)))
  countReactOnK = list(map(int, countReactOnK))
  countReactOnK = sum(countReactOnK)
  return countReactOnK

def get_posts_type(message):
  if any(word in message.message for word in AD_STOPWORDS):
    adsOrContent = 'Реклама/Ивент/Вакансии'
  else:
    adsOrContent = 'Контент'
  return adsOrContent

def paste_to_Excel(ws, postsQty, i, postsDateExl, messageTextExl, chat, countForOnP, countReactOnP, countComOnP, countReactOnKFinal, likeDislikeExcel, countVotes, viewsExl, adsOrContent):
  ws['A' + str(postsQty)] = postsDateExl
  ws['B' + str(postsQty)] = messageTextExl
  ws['C' + str(postsQty)] = chat + "/" + str(i)
  ws['D' + str(postsQty)] = str(countForOnP)
  ws['E' + str(postsQty)] = str(countReactOnP)
  ws['F' + str(postsQty)] = str(countComOnP)
  ws['G' + str(postsQty)] = str(countReactOnKFinal)
  ws['H' + str(postsQty)] = likeDislikeExcel
  ws['I' + str(postsQty)] = countVotes
  ws['J' + str(postsQty)] = viewsExl
  ws['K' + str(postsQty)] = adsOrContent

async def get_data_to_google_docs(service, SPREADSHEET_ID, sheet_name, lastRowId, postsQty, excelPasteData):
  # Начинаем запись с последней строки и прибавляем номер поста
  rangeExcel = sheet_name +  "!A" + str(lastRowId + 1) + ":K" + str(lastRowId + 1 + postsQty)
  values = service.spreadsheets().values().batchUpdate(
    spreadsheetId=SPREADSHEET_ID,
    body={
      "valueInputOption": "USER_ENTERED",
      "data": [
          {"range": rangeExcel,
          "majorDimension": "ROWS",
          "values": excelPasteData}
      ]
    }).execute()

async def save_Excel(msg, chat_name_url, start_date, end_date, wb):
  # Очищаем название файла от косых черт и сохраняем с расширением *.xlsx
  chatNameVerified = str(chat_name_url[0])
  chatNameVerified = chatNameVerified.split("/")[0]
  documentName = 'Архив постов "' + chatNameVerified + '" (' + str(start_date) + ' - ' + str(end_date) + ').xlsx'
  wb.save(documentName)
  await bot.send_document(msg.from_user.id, open(documentName, 'rb'), caption = "Архив постов для " + str(chat_name_url[0]) + " готов!")
  try: os.remove(documentName)
  except OSError: await bot.send_message(msg.from_user.id, "Почему-то не получилось удалить файл с сервера", reply_markup=bot_keyboard.start)

if __name__ == '__main__':
  print('Монстр пчелы запущен!')

executor.start_polling(dp)
